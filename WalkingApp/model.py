import openpyxl
from datetime import datetime, timedelta


##################
# Event Handlers #
##################
def _open_excel(data):
    """opens the Excel file
    :param data: data according to whether data_only are needed"""
    try:
        file = "walking_data.xlsx"
        load_file = openpyxl.load_workbook(file, data_only=data)
        return load_file
    except FileNotFoundError:
        template = "new_template.xlsx"
        load_file = openpyxl.load_workbook(template, data_only=data)
        return load_file


def get_recent_walks():
    """gets four last walks from Excel and returns them"""
    wb = _open_excel(True)
    ws = wb["Sheet1"]

    data = []
    if ws.max_row >= 2:
        num_records = min(ws.max_row - 1, 4)
        for num in range(ws.max_row, ws.max_row - num_records, -1):
            date = ws[f'a{num}'].value if type(ws[f'a{num}'].value) is str else ws[f'a{num}'].value.strftime('%d/%m/%Y')
            kms = ws[f'b{num}'].value
            data.append((date, kms))

    wb.close()
    return data


def save_to_excel(date, kms: float, time, kcal: int, steps: int):
    """gets checked data as params and save them into Excel file
    :param date: date value in required format
    :param kms: kilometres, float value
    :param time: time value in required format
    :param kcal: amount of kcal, int value
    :param steps: amount of steps, int value"""
    wb = _open_excel(False)
    ws = wb["Sheet1"]

    insert_into_excel = [date,
                         kms,
                         time,
                         kcal,
                         steps]
    ws.append(insert_into_excel)

    wb.save("walking_data.xlsx")
    wb.close()


def calculate_statistics():
    """receives from Excel file data, calculates statistics and returns them"""
    workbook = _open_excel(True)
    worksheet = workbook["Sheet1"]

    column_km = worksheet["B"]
    column_kcal = worksheet["D"]
    column_steps = worksheet["E"]
    total_time = timedelta()

    total_km = sum(float(cell.value) for cell in column_km if isinstance(cell.value, (int, float)))

    for row in worksheet.iter_rows(min_row=2, max_col=3, values_only=True):
        time = datetime.strptime(str(row[2]), "%H:%M:%S").time()
        time_timedelta = timedelta(
            hours=time.hour,
            minutes=time.minute,
            seconds=time.second
        )
        total_time += time_timedelta

    total_kcal = sum(int(cell.value) for cell in column_kcal if isinstance(cell.value, (int, float)))

    total_steps = sum(int(cell.value) for cell in column_steps if isinstance(cell.value, (int, float)))

    workbook.close()
    return total_km, total_time, total_kcal, total_steps


def set_selected_date():
    pass


def get_selected_date():
    pass


###########
#  View   #
###########
selected_date = None
