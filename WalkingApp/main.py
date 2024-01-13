import openpyxl
import flet as ft
from datetime import datetime, timedelta
import re
import dialogs

pattern_hours_minutes = r'^([0-9]|1[0-2]|2[0-3]):[0-5][0-9]$'
pattern_hours = r'^(1?[0-9]|2[0-3])$'


# FUNCTIONS THAT WORKS WITH EXCEL
def open_excel(data):
    try:
        file = "walking_data.xlsx"
        load_file = openpyxl.load_workbook(file, data_only=data)
        return load_file
    except FileNotFoundError:
        template = "new_template.xlsx"
        load_file = openpyxl.load_workbook(template, data_only=data)
        return load_file


def get_recent_walks():
    wb = open_excel(True)
    ws = wb["Sheet1"]

    data = []
    if ws.max_row >= 2:
        num_records = min(ws.max_row - 1, 4)
        for num in range(ws.max_row, ws.max_row - num_records, -1):
            date = ws[f'a{num}'].value if type(ws[f'a{num}'].value) is str else ws[f'a{num}'].value.strftime('%d/%m/%Y')
            kms = ws[f'b{num}'].value
            data.append((date, kms))

    wb.close()
    return data


def save_to_excel(date, kms: float, time, kcal: int, steps: int):
    wb = open_excel(False)
    ws = wb["Sheet1"]

    insert_into_excel = [date,
                         kms,
                         time,
                         kcal,
                         steps]
    ws.append(insert_into_excel)

    wb.save("walking_data.xlsx")
    wb.close()


def statistics():
    workbook = open_excel(True)
    worksheet = workbook["Sheet1"]

    column_km = worksheet["B"]
    column_kcal = worksheet["D"]
    column_steps = worksheet["E"]
    total_time = timedelta()

    total_km = sum(float(cell.value) for cell in column_km if isinstance(cell.value, (int, float)))

    for row in worksheet.iter_rows(min_row=2, max_col=3, values_only=True):
        time = datetime.strptime(str(row[2]), "%H:%M:%S").time()
        time_timedelta = timedelta(
            hours=time.hour,
            minutes=time.minute,
            seconds=time.second
        )
        total_time += time_timedelta

    total_kcal = sum(int(cell.value) for cell in column_kcal if isinstance(cell.value, (int, float)))

    total_steps = sum(int(cell.value) for cell in column_steps if isinstance(cell.value, (int, float)))

    workbook.close()
    return total_km, total_time, total_kcal, total_steps


# MAIN BODY OF THE APP WRAP IN FUNCTION MAIN
def main(page: ft.Page):
    page.title = "WalkingApp"
    page.padding = 0
    page.window_width = 400
    page.window_height = 850
    page.bgcolor = ft.colors.BLUE_100
    page.window_resizable = False
    page.window_maximizable = False
    page.theme_mode = ft.ThemeMode.LIGHT

    def window_event(e):
        if e.data == "close" or e.name == "click":
            dialogs.show_confirm_dialog(page)

    page.window_prevent_close = True
    page.on_window_event = window_event

    def open_statistics(_):
        page.go("/statistics")

    def fill_recent_walks_table():
        walks_data = get_recent_walks()
        recent_walks = walks_data if walks_data else [("Žádné záznamy", "")]
        data_table.rows = [
            ft.DataRow(
                [ft.DataCell(ft.Text(date)), ft.DataCell(ft.Text(kms))]
            ) for date, kms in recent_walks]
        page.update()

    def pick_date(_):
        picked_date.value = "Budeš přidávat aktivitu ze dne {}".format(date_picker.value.strftime("%d/%m/%y"))
        page.update()

    def open_google_maps(_):
        page.launch_url("https://mapy.cz/")

    def save_time_entry(walked_time_entry_value):
        if re.search(pattern_hours_minutes, walked_time_entry_value):
            time_format_to_save = f"{walked_time_entry_value}:00"
            return time_format_to_save
        elif re.search(pattern_hours, walked_time_entry_value):
            time_format_to_save = f"{walked_time_entry_value}:00:00"
            return time_format_to_save
        else:
            dialogs.show_wrong_time_dialog(page)
            return None

    def save_clicked(_):
        if not all([walked_time_entry.value, walked_kms_entry.value, walked_kcal_entry.value,
                    walked_steps_entry.value]):
            dialogs.show_incomplete_dialog(page)
            return

        if not date_picker.value:
            page.go("/")
            dialogs.show_no_date_picked_dialog(page)
            return

        date = date_picker.value.strftime("%d/%m/%y")
        kms = float(walked_kms_entry.value)
        time = save_time_entry(walked_time_entry.value)
        kcal = int(walked_kcal_entry.value)
        steps = int(walked_steps_entry.value)

        if time is None:
            return

        save_to_excel(date, kms, time, kcal, steps)

        dialogs.show_success_dialogue(page, fill_recent_walks_table)

        walked_kms_entry.value = ""
        walked_time_entry.value = ""
        walked_kcal_entry.value = ""
        walked_steps_entry.value = ""
        page.update()

    # CONTROLS OF THE APP - BUTTONS, TEXTS

    welcome_txt = ft.Text(value="\nVítej ve WalkingApp!\n",
                          color=ft.colors.INDIGO,
                          style=ft.TextThemeStyle.DISPLAY_MEDIUM,
                          text_align=ft.TextAlign.CENTER)

    walked_kms_entry = ft.TextField(label="Kolik jsi ušel?",
                                    hint_text="km.m",
                                    width=160,
                                    border_radius=0,
                                    input_filter=ft.InputFilter(allow=True, regex_string=r"[0-9.]",
                                                                replacement_string=""),
                                    keyboard_type=ft.KeyboardType.NUMBER)
    walked_time_entry = ft.TextField(label="Za jak dlouho?",
                                     hint_text="hodiny:minuty",
                                     width=160,
                                     border_radius=0,
                                     input_filter=ft.InputFilter(allow=True, regex_string=r"[0-9:]",
                                                                 replacement_string=""))
    walked_kcal_entry = ft.TextField(label="Kolik kalorií?",
                                     width=160,
                                     border_radius=0,
                                     input_filter=ft.InputFilter(allow=True, regex_string=r"[0-9]",
                                                                 replacement_string=""),
                                     keyboard_type=ft.KeyboardType.NUMBER)
    walked_steps_entry = ft.TextField(label="A kolik kroků?",
                                      width=160,
                                      border_radius=0,
                                      input_filter=ft.InputFilter(allow=True, regex_string=r"[0-9]",
                                                                  replacement_string=""),
                                      keyboard_type=ft.KeyboardType.NUMBER)

    date_picker = ft.DatePicker(on_change=pick_date,
                                first_date=datetime(2023, 10, 1),
                                last_date=datetime(2030, 12, 31)
                                )

    date_button = ft.ElevatedButton("Vyber datum",
                                    icon=ft.icons.CALENDAR_MONTH_ROUNDED,
                                    on_click=lambda _: date_picker.pick_date(),
                                    )
    picked_date = ft.Text(value="Vyber datum",
                          text_align=ft.TextAlign.CENTER,
                          color=ft.colors.WHITE54)

    data_table = ft.DataTable(
        bgcolor=ft.colors.WHITE54,
        columns=[
            ft.DataColumn(ft.Text("Datum")),
            ft.DataColumn(ft.Text("Kilometry")),
        ],
        rows=[]
    )

    new_record_button = ft.FilledButton(text="Přidej nový záznam",
                                        on_click=lambda _: page.go("/new"))

    exit_button = ft.ElevatedButton(text="Konec",
                                    style=ft.ButtonStyle(
                                        shape=ft.ContinuousRectangleBorder(radius=30)),
                                    on_click=window_event)

    save_button = ft.ElevatedButton(text="Uložit",
                                    on_click=save_clicked)

    show_statistics_button = ft.ElevatedButton(text="Ukaž statistiky",
                                               on_click=open_statistics)

    fill_recent_walks_table()
    page.overlay.append(date_picker)

    open_maps = ft.Chip(
        label=ft.Text("Nápad na trasu"),
        leading=ft.Icon(ft.icons.MAP_SHARP),
        on_click=open_google_maps,
    )

    # GUI OF THE APP

    def views(_):
        page.views.clear()
        page.views.append(
            ft.View(
                "/",
                bgcolor=ft.colors.BLUE_100,
                padding=0,
                controls=[ft.Stack(
                    [
                        ft.Image(
                            src="https://picsum.photos/id/651/400/800",
                            width=page.window_width,
                            height=page.window_height,
                            fit=ft.ImageFit.FILL
                        ),
                        ft.Container(left=65,
                                     height=200,
                                     width=270,
                                     content=welcome_txt),
                        ft.Container(right=100,
                                     top=280,
                                     width=180,
                                     height=35,
                                     content=date_button),
                        ft.Container(right=90,
                                     top=325,
                                     height=50,
                                     width=200,
                                     content=picked_date),
                        ft.Container(right=90,
                                     top=420,
                                     width=200,
                                     height=30,
                                     content=new_record_button),
                        ft.Container(right=90,
                                     bottom=195,
                                     width=200,
                                     height=50,
                                     content=open_maps),
                        ft.Container(right=5,
                                     bottom=80,
                                     width=100,
                                     height=25,
                                     content=exit_button),
                    ],
                    width=page.window_width,
                    height=page.window_height)],
            )
        )
        if page.route == "/new" or page.route == "/statistics":
            page.views.append(
                ft.View(
                    route="/new",
                    bgcolor=ft.colors.BLUE_100,
                    padding=0,
                    appbar=ft.AppBar(title=ft.Text("Nový záznam"),
                                     bgcolor=ft.colors.BLUE_100),
                    controls=[
                        ft.Stack(controls=[
                            ft.Image(
                                src="https://picsum.photos/id/651/400/800",
                                width=page.window_width,
                                height=page.window_height,
                                fit=ft.ImageFit.FILL),
                            ft.Container(content=walked_kms_entry,
                                         top=30,
                                         left=25,
                                         bgcolor=ft.colors.BLUE_50),
                            ft.Container(content=walked_time_entry,
                                         top=30,
                                         left=200,
                                         bgcolor=ft.colors.BLUE_50),
                            ft.Container(content=walked_kcal_entry,
                                         top=110,
                                         left=25,
                                         bgcolor=ft.colors.BLUE_50),
                            ft.Container(content=walked_steps_entry,
                                         top=110,
                                         left=200,
                                         bgcolor=ft.colors.BLUE_50),
                            ft.Container(content=save_button,
                                         top=200,
                                         left=20),
                            ft.Container(content=show_statistics_button,
                                         left=20,
                                         top=250),
                            ft.Container(content=data_table,
                                         left=70,
                                         top=350),
                            ft.Container(content=exit_button,
                                         right=5,
                                         bottom=80,
                                         width=100,
                                         height=25),

                        ],
                            width=page.window_width,
                            height=page.window_height - 70)]
                )
            )

        if page.route == "/statistics":
            total_km, total_time, total_kcal, total_steps = statistics()
            page.views.append(
                ft.View(
                    "/statistics",
                    bgcolor=ft.colors.BLUE_100,
                    appbar=ft.AppBar(title=ft.Text("Statistiky"),
                                     bgcolor=ft.colors.BLUE_100),
                    padding=0,
                    controls=[
                        ft.Stack(controls=[
                            ft.Image(
                                src="https://picsum.photos/id/651/400/800",
                                width=page.window_width,
                                height=page.window_height,
                                fit=ft.ImageFit.FILL),
                            ft.Container(content=ft.Text("{} Km".format(round(total_km))),
                                         margin=10,
                                         padding=10,
                                         alignment=ft.alignment.center,
                                         bgcolor=ft.colors.AMBER,
                                         border_radius=90,
                                         width=175,
                                         height=100,
                                         left=100,
                                         top=20),
                            ft.Container(content=ft.Text("{} hod.".format(total_time)),
                                         margin=10,
                                         padding=10,
                                         alignment=ft.alignment.center,
                                         bgcolor=ft.colors.GREEN_200,
                                         border_radius=90,
                                         width=175,
                                         height=100,
                                         left=100,
                                         top=150),
                            ft.Container(content=ft.Text("{} Kcal".format(total_kcal)),
                                         margin=10,
                                         padding=10,
                                         alignment=ft.alignment.center,
                                         bgcolor=ft.colors.CYAN_200,
                                         border_radius=90,
                                         width=175,
                                         height=100,
                                         left=100,
                                         top=280),
                            ft.Container(content=ft.Text("{} kroků".format(total_steps)),
                                         margin=10,
                                         padding=10,
                                         alignment=ft.alignment.center,
                                         bgcolor=ft.colors.RED_200,
                                         border_radius=90,
                                         width=175,
                                         height=100,
                                         left=100,
                                         top=410),
                            ft.Container(content=exit_button,
                                         right=5,
                                         bottom=80,
                                         width=100,
                                         height=25),
                        ],
                            width=page.window_width,
                            height=page.window_height - 70)]
                )
            )
        page.update()

    def view_pop(_):
        page.views.pop()
        top_view = page.views[-1]
        page.go(top_view.route)

    page.on_route_change = views
    page.on_view_pop = view_pop
    page.go(page.route)


# START OF THE APP

ft.app(target=main)
